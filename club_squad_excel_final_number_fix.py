
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import tempfile

# Function to process squad data
def process_squad(raw_text, club_name):
    known_positions = {"GK", "CB", "LB", "RB", "DM", "CM", "RM", "LM", "AM", "LW", "RW", "SS", "CF"}
    players = []
    lines = raw_text.strip().split("\n")

    i = 0
    while i < len(lines):
        number_line = lines[i].strip()
        if not number_line.isdigit():
            i += 1
            continue

        number = int(number_line)  # now store as real number
        if i + 1 >= len(lines):
            break

        name_and_positions = lines[i + 1].strip().split()
        name_parts = []
        position_parts = []

        for part in name_and_positions:
            if part in known_positions:
                position_parts.append(part)
            else:
                name_parts.append(part)

        name = " ".join(name_parts)
        players.append((number, name, position_parts))
        i += 2

    wb = Workbook()
    ws = wb.active
    ws.title = "Squad"

    ws.merge_cells('A1:O1')
    ws['A1'] = f"{club_name} Squad List"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Number", "Name", "GK", "CB", "LB", "RB", "DM", "CM", "RM", "LM", "AM", "LW", "RW", "SS", "CF"]
    ws.append(headers)

    fills = {
        "GK": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "CB": PatternFill(start_color="FFD580", end_color="FFD580", fill_type="solid"),
        "LB": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "RB": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "LM": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
        "LW": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
        "DM": PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid"),
        "CM": PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid"),
        "AM": PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid"),
        "RM": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        "RW": PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"),
        "SS": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
        "CF": PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"),
    }

    for idx, (number, name, positions) in enumerate(players, start=3):
        row_data = {pos: "" for pos in headers}
        row_data["Number"] = number
        row_data["Name"] = name

        for pos in positions:
            if pos in headers:
                row_data[pos] = pos

        row = [row_data[col] for col in headers]
        ws.append(row)

        for col_idx, pos in enumerate(headers[2:], start=3):
            if row_data[pos] != "":
                cell = ws.cell(row=idx, column=col_idx)
                if pos in fills:
                    cell.fill = fills[pos]

    for idx, cell in enumerate(ws[2], start=1):
        cell.font = Font(bold=True)
        if idx == 1:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif idx == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        if i == 1:
            ws.column_dimensions[column_letter].width = 8  # nice fixed width for numbers
        elif i == 2:
            ws.column_dimensions[column_letter].width = max_length + 4  # name wider
        else:
            ws.column_dimensions[column_letter].width = max_length + 2

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    return temp_file.name

# Streamlit App
st.title("⚽ Club Squad Excel Generator (Final Number Fix)")
raw_input_text = st.text_area("Paste your squad list here:", height=400)
club = st.text_input("Enter club name (e.g., Brighton FC):")

if st.button("Generate Excel"):
    if raw_input_text and club:
        with st.spinner('Generating your squad Excel...'):
            file_path = process_squad(raw_input_text, club)
        st.success('Squad Excel ready! 🎉')
        with open(file_path, "rb") as f:
            st.download_button("Download Squad Excel", f, file_name=f"{club.replace(' ', '_').lower()}_squad.xlsx")
    else:
        st.error("Please provide both squad list and club name.")
